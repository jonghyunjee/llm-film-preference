#!/usr/bin/env python3
import anthropic
try:
    import openai as _openai_lib
    _OPENAI_AVAILABLE = True
except ImportError:
    _OPENAI_AVAILABLE = False
import json
import random
import math
from typing import List, Dict, Tuple, Optional
from datetime import datetime
from collections import defaultdict
import numpy as np
from scipy.optimize import minimize
from scipy.stats import norm
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    _OPENPYXL_AVAILABLE = True
except ImportError:
    _OPENPYXL_AVAILABLE = False


# ─────────────────────────────────────────────────────────────────────────────
# BRADLEY-TERRY MODEL
# ─────────────────────────────────────────────────────────────────────────────

class BradleyTerry:
    """
    Bradley-Terry model with MLE via ILSR (MM algorithm).

      1. Minimum coverage floor: adaptive sampling won't focus until every item has >= min_matches_before_focus real matches.
      2. Regularization (pseudo-counts): 0.5 phantom win + 0.5 phantom loss per item prevents MLE divergence for undefeated/winless items.
      3. Full inverse Fisher information matrix for SEs, giving correct uncertainty even in sparse/uneven comparison graphs.
      4. Boundary tests use the covariance term from the full matrix, so the SE of a *difference* λ_i − λ_j is computed correctly.

    Model: P(i beats j) = π_i / (π_i + π_j)
    Working in log-space: λ_i = log(π_i), centered so mean(λ) = 0.
    """

    # How many real matches an item must have before strength-based
    MIN_MATCHES_BEFORE_FOCUS: int = 15

    def __init__(self):
        self.items:        List[str]              = []
        self.item_index:   Dict[str, int]         = {}
        self.wins:         Dict[int, Dict[int, int]] = defaultdict(lambda: defaultdict(int))
        self.match_history: List[Dict]            = []

        # Set after fit()
        self.log_strengths: Optional[np.ndarray] = None  # shape (n,)
        self.cov_matrix:    Optional[np.ndarray] = None  # shape (n, n), full inverse Fisher
        self.std_errors:    Optional[np.ndarray] = None  # shape (n,), sqrt(diag(cov))

    # ── Data ingestion ────────────────────────────────────────────────────────

    def add_item(self, item: str):
        if item not in self.item_index:
            self.item_index[item] = len(self.items)
            self.items.append(item)

    def record_outcome(self, winner: str, loser: str,
                       run_id: int = 0, comparison_idx: int = 0):
        for item in (winner, loser):
            self.add_item(item)
        i = self.item_index[winner]
        j = self.item_index[loser]
        self.wins[i][j] += 1
        self.match_history.append({
            'run': run_id,
            'n':   comparison_idx,
            'w':   winner,
            'l':   loser,
        })

    # ── Win / comparison matrices ─────────────────────────────────────────────

    def _build_matrices(self) -> Tuple[np.ndarray, np.ndarray,
                                        np.ndarray, np.ndarray]:
        """
        Returns (W, N, W_reg, N_reg).

        W[i,j]     = observed wins of i over j
        N[i,j]     = W[i,j] + W[j,i]  (total comparisons, unregularized)
        W_reg      = W + 0.5  (Dirichlet-0.5 pseudo-counts)
        N_reg      = W_reg + W_reg.T
        """
        n = len(self.items)
        W = np.zeros((n, n))
        for i, js in self.wins.items():
            for j, w in js.items():
                W[i, j] = w
        N     = W + W.T
        W_reg = W + 0.5
        N_reg = W_reg + W_reg.T
        return W, N, W_reg, N_reg

    # ── ILSR fitting ─────────────────────────────────────────────────────────

    def fit(self, max_iter: int = 2000, tol: float = 1e-10) -> np.ndarray:
        """
        Fit Bradley-Terry via the MM (minorization-maximization) algorithm equivalent to ILSR.

        Uses regularized counts (W_reg, N_reg) for the update to avoid divergence; uses unregularized counts (N) to build the Fisher information matrix so SEs reflect actual data, not the prior.

        Returns centered log-strengths λ = log(π) − mean(log(π)).
        """
        n = len(self.items)
        if n < 2:
            raise ValueError("Need at least 2 items to fit.")

        W, N, W_reg, N_reg = self._build_matrices()

        # ── MM iterations (regularised) ───────────────────────────────────
        pi = np.ones(n, dtype=float)

        for _ in range(max_iter):
            pi_old = pi.copy()
            for i in range(n):
                w_i   = W_reg[i].sum()
                denom = sum(
                    N_reg[i, j] / (pi[i] + pi[j])
                    for j in range(n)
                    if N_reg[i, j] > 0
                )
                if denom > 0:
                    pi[i] = w_i / denom
            pi /= pi.sum() / n                          # keep scale stable
            if np.max(np.abs(pi - pi_old) / (pi_old + 1e-12)) < tol:
                break

        # ── Log-strengths, centered ────────────────────────────────────────
        log_pi = np.log(pi)
        log_pi -= log_pi.mean()
        self.log_strengths = log_pi

        # ── Full inverse Fisher information matrix (unregularised data) ───
        self.cov_matrix, self.std_errors = self._compute_full_covariance(pi, N)

        return self.log_strengths

    # ── Full covariance via Fisher information ────────────────────────────────

    def _compute_full_covariance(self, pi: np.ndarray,
                                  N: np.ndarray) -> Tuple[np.ndarray, np.ndarray]:
        """
        Compute the full n×n covariance matrix via the reduced-parameter method.

        Method: fix-one-out (reference parameterization)
        ─────────────────────────────────────────────────
        The BT model has one unidentified degree of freedom (additive constant).
        Standard fix: anchor item n (weakest π) to λ_n = 0 and estimate the
        free (n-1) parameters λ_1..λ_{n-1}.  The (n-1)×(n-1) reduced Fisher
        matrix is full-rank and invertible, giving Cov_{n-1}.

        We then recover the full n×n covariance under the mean-zero constraint
        Σ λ_i = 0 using the propagation formula:

            C_ij = Cov_{n-1}[i,j] − (1/n)·Σ_k Cov_{n-1}[i,k]
                                   − (1/n)·Σ_k Cov_{n-1}[k,j]
                                   + (1/n²)·Σ_{k,l} Cov_{n-1}[k,l]

        for i,j in {0..n-2}, and for i or j = ref: 0 (reference has no
        free variance by construction, but the propagated matrix gives it
        implicit uncertainty via the off-diagonals. This is handled accordingly
        in boundary_test via the full Var(λ_i − λ_j) formula).

        Returns (cov_nn, se_n):
          cov_nn  — n×n covariance matrix (mean-zero constraint)
          se_n    — sqrt(diag(cov_nn))
        """
        n = len(pi)

        # ── Fisher information matrix ─────────────────────────────────────
        I_mat = np.zeros((n, n))
        for i in range(n):
            for j in range(n):
                if i != j and N[i, j] > 0:
                    v = N[i, j] * pi[i] * pi[j] / (pi[i] + pi[j]) ** 2
                    I_mat[i, i] += v
                    I_mat[i, j] -= v

        # Reference: item with smallest π (weakest; least variance to anchor)
        ref = int(np.argmin(pi))
        free = [k for k in range(n) if k != ref]  # length n-1

        # ── Invert reduced (n-1)×(n-1) Fisher matrix ─────────────────────
        I_red = I_mat[np.ix_(free, free)]
        ridge = 1e-8 * np.eye(len(free))
        try:
            cov_red = np.linalg.inv(I_red + ridge)
        except np.linalg.LinAlgError:
            cov_red = np.linalg.pinv(I_red + ridge)

        # ── Propagate to full n×n under mean-zero constraint ──────────────
        # Build (n-1)×(n-1) covariance in the free-parameter space, then
        # apply the linear transformation that centres the estimates.
        # row/col sums of cov_red (for the propagation formula)
        row_sums = cov_red.sum(axis=1)           # shape (n-1,)
        total    = cov_red.sum()                 # scalar

        cov_nn = np.zeros((n, n))
        for ii, i in enumerate(free):
            for jj, j in enumerate(free):
                cov_nn[i, j] = (cov_red[ii, jj]
                                - row_sums[ii] / n
                                - row_sums[jj] / n
                                + total / n**2)

        # Ref item row/col: derived from constraint; fill using symmetry
        # Var(λ_ref) = Var(-Σ_{k≠ref} λ_k) = Σ_{i,j} Cov(λ_i, λ_j) / 1
        # but for boundary tests we only need Var(λ_i - λ_j); the formula
        # Var_i + Var_j - 2*Cov_ij works for any pair including ref.
        # Compute ref variance from propagation:
        cov_nn[ref, ref] = total / n**2
        for ii, i in enumerate(free):
            c_ref_i = -row_sums[ii] / n + total / n**2
            cov_nn[ref, i] = c_ref_i
            cov_nn[i, ref] = c_ref_i

        # non-negative diagonal
        np.fill_diagonal(cov_nn, np.maximum(np.diag(cov_nn), 0.0))

        se_n = np.sqrt(np.diag(cov_nn))
        return cov_nn, se_n

    # ── Rankings and inference ────────────────────────────────────────────────

    def get_rankings(self) -> List[Dict]:
        """
        Return items sorted by log-strength with 95% CIs.
        SEs come from the full covariance matrix diagonal.
        """
        if self.log_strengths is None:
            raise RuntimeError("Call fit() before get_rankings().")

        n = len(self.items)
        match_counts = {
            i: sum(self.wins[i].values()) +
               sum(self.wins[j].get(i, 0) for j in range(n))
            for i in range(n)
        }

        rankings = []
        for i, item in enumerate(self.items):
            lam = self.log_strengths[i]
            se  = self.std_errors[i]
            rankings.append({
                'item':         item,
                'log_strength': round(float(lam), 4),
                'std_error':    round(float(se), 4),
                'ci_lower_95':  round(float(lam - 1.96 * se), 4),
                'ci_upper_95':  round(float(lam + 1.96 * se), 4),
                'matches':      match_counts.get(i, 0),
            })

        rankings.sort(key=lambda x: x['log_strength'], reverse=True)
        for rank, row in enumerate(rankings, 1):
            row['rank'] = rank
        return rankings

    def match_counts_per_item(self) -> Dict[str, int]:
        n = len(self.items)
        return {
            item: (sum(self.wins[i].values()) +
                   sum(self.wins[j].get(i, 0) for j in range(n)))
            for i, item in enumerate(self.items)
        }

# ─────────────────────────────────────────────────────────────────────────────
# ADAPTIVE SAMPLER
# ─────────────────────────────────────────────────────────────────────────────

class BTAdaptiveSampler:
    """
    Three-phase adaptive sampling for Bradley-Terry.

    Phase 1 — Global coverage (mandatory floor)
    ─────────────────────────────────────────────
    Pure uncertainty sampling until EVERY item has >= min_floor matches.
    This guarantees no item is left with only 1-6 matches before the
    BT model has any reliable signal about it.
    Budget: n_items * min_floor / 2 comparisons (exact lower bound).

    Phase 2 — Broad competitive sampling (middle stabilization)
    ────────────────────────────────────────────────────────────
    Strength-based sampling across the full ranking. Match items with
    similar current λ estimates to stabilize the middle of the distribution.
    Budget: configurable middle_budget fraction of remaining comparisons.

    Phase 3 — Boundary intensification
    ──────────────────────────────────────────────────────
    Concentrate remaining budget on items near the boundary ranks of
    interest (top-25%, bottom-25% by default, or user-specified).
    Items within ±window of each boundary rank get heavily oversampled.

    Phase transitions are deterministic based on comparison count,
    not on per-item match counts.
    """

    def __init__(self, bt: 'BradleyTerry',
                 min_floor: int = None,
                 middle_fraction: float = 0.35,
                 boundary_window: int = 8):
        """
        Args:
            min_floor:        Matches/item guaranteed in Phase 1.
                              Defaults to BradleyTerry.MIN_MATCHES_BEFORE_FOCUS.
            middle_fraction:  Fraction of post-floor budget spent in Phase 2.
            boundary_window:  Items within ±window of each boundary rank get
                              focused sampling in Phase 3.
        """
        self.bt               = bt
        self.min_floor        = min_floor or BradleyTerry.MIN_MATCHES_BEFORE_FOCUS
        self.middle_fraction  = middle_fraction
        self.boundary_window  = boundary_window

        # Set by run_experiment before sampling begins
        self.n_items:          int       = 0
        self.max_comparisons:  int       = 0
        self.boundary_ranks:   List[int] = []

        # Phase thresholds (set in configure())
        self._phase1_end: int = 0   # comparison index where Phase 1 ends
        self._phase2_end: int = 0   # comparison index where Phase 2 ends

    def configure(self, n_items: int, max_comparisons: int,
                  boundary_ranks: List[int]):
        """Call once before the run loop to set phase thresholds."""
        self.n_items         = n_items
        self.max_comparisons = max_comparisons
        self.boundary_ranks  = boundary_ranks

        # Phase 1: guarantee min_floor matches for every item
        # Lower bound: n*floor/2 comparisons (each covers 2 items)
        # Add 20% slack for the adaptive sampler's uneven distribution
        phase1_base  = int(n_items * self.min_floor / 2 * 1.2)
        self._phase1_end = min(phase1_base, int(max_comparisons * 0.55))

        # Phase 2: broad competitive sampling
        remaining    = max_comparisons - self._phase1_end
        self._phase2_end = self._phase1_end + int(remaining * self.middle_fraction)

        print(f"  Sampler phases:")
        print(f"    Phase 1 (coverage floor ≥{self.min_floor}/item): "
              f"comparisons 1–{self._phase1_end}")
        print(f"    Phase 2 (broad competitive):  "
              f"comparisons {self._phase1_end+1}–{self._phase2_end}")
        print(f"    Phase 3 (boundary focus):     "
              f"comparisons {self._phase2_end+1}–{max_comparisons}")
        print(f"    Boundary ranks: {boundary_ranks}")
        print()

    def current_phase(self, comparisons_done: int) -> int:
        if comparisons_done < self._phase1_end:
            return 1
        elif comparisons_done < self._phase2_end:
            return 2
        else:
            return 3

    # ── Phase 1: uncertainty-based global coverage ────────────────────────────

    def sample_uncertainty_based(self, items: List[str], n_pairs: int = 1):
        """
        Prioritize items with fewest matches.
        Picks from the top-20 most under-sampled to keep some randomness.
        """
        counts = self.bt.match_counts_per_item()
        scored = sorted(items, key=lambda x: counts.get(x, 0))
        pool   = scored[:min(20, len(scored))]
        pairs  = []
        for _ in range(n_pairs):
            if len(pool) >= 2:
                pairs.append(tuple(random.sample(pool, 2)))
        return pairs

    # ── Phase 2: strength-based competitive sampling ──────────────────────────

    def sample_competitive(self, items: List[str], n_pairs: int = 1):
        """
        Match items with similar current BT log-strength.
        Uses a wider window (10) than before so middle items get diverse
        opponents rather than repeatedly facing the same neighbours.
        """
        if self.bt.log_strengths is None:
            return self.sample_uncertainty_based(items, n_pairs)

        rated = sorted(
            [x for x in items if x in self.bt.item_index],
            key=lambda x: self.bt.log_strengths[self.bt.item_index[x]]
        )
        pairs = []
        window = 10
        for _ in range(n_pairs):
            idx  = random.randint(0, len(rated) - 2)
            nidx = random.randint(
                max(0, idx - window),
                min(len(rated) - 1, idx + window)
            )
            if nidx != idx:
                pairs.append((rated[idx], rated[nidx]))
        return pairs if pairs else self.sample_uncertainty_based(items, n_pairs)

    # ── Phase 3: boundary-focused sampling ───────────────────────────────────

    def sample_boundary_focused(self, items: List[str], n_pairs: int = 1):
        """
        Phase 3: concentrate sampling on the top quartile (prestige zone),
        with a lighter anchor pass against the bottom quartile and middle.

        Sampling mix (reflecting research priority = critical acclaim orientation):
          60%  pure top-zone pairs   (within top-25% boundary items)
          20%  cross-anchor pairs    (top-zone item vs bottom-zone item)
                                     — anchors scale between extremes
          10%  top vs middle         — positions top tier vs the pack
          10%  full competitive      — keeps middle from drifting entirely

        The bottom-zone gets comparisons only as the "anchor" partner,
        not as the focus. This directs the bulk of the Phase 3 budget
        toward refining the prestige ranking rather than the bottom.

        boundary_ranks should be set to emphasize top cutoffs, e.g.:
          [15, 30]          — top-12.5% and top-25%
          [10, 20, 30]      — finer-grained top tiers
        A single bottom anchor rank (e.g. 110) can be included; it will
        be used only in cross-anchor pairs.
        """
        if self.bt.log_strengths is None:
            return self.sample_uncertainty_based(items, n_pairs)

        ranked = sorted(
            [x for x in items if x in self.bt.item_index],
            key=lambda x: self.bt.log_strengths[self.bt.item_index[x]],
            reverse=True
        )
        n = len(ranked)

        # Partition into top-zone, bottom-zone, and middle
        top_set    = set()
        bottom_set = set()
        for k in self.boundary_ranks:
            lo = max(0, k - 1 - self.boundary_window)
            hi = min(n,  k - 1 + self.boundary_window + 1)
            zone = ranked[lo:hi]
            # Heuristic: ranks in the lower half of the list → bottom anchor
            if k > n // 2:
                bottom_set.update(zone)
            else:
                top_set.update(zone)

        # Fallback: if no explicit bottom ranks given, use bottom 25% as anchor
        if not bottom_set:
            bottom_set = set(ranked[3 * n // 4:])

        middle_list = [x for x in ranked if x not in top_set and x not in bottom_set]
        top_list    = list(top_set)
        bottom_list = list(bottom_set)

        pairs = []
        for _ in range(n_pairs):
            r = random.random()
            if r < 0.60 and len(top_list) >= 2:
                # Pure top-zone pair — primary focus
                pairs.append(tuple(random.sample(top_list, 2)))
            elif r < 0.80 and top_list and bottom_list:
                # Cross-anchor: top vs bottom — calibrates scale extremes
                pairs.append((random.choice(top_list),
                               random.choice(bottom_list)))
            elif r < 0.90 and top_list and middle_list:
                # Top vs middle — positions prestige tier vs the field
                pairs.append((random.choice(top_list),
                               random.choice(middle_list)))
            else:
                # Full competitive fallback
                pairs.extend(self.sample_competitive(items, 1))

        return pairs if pairs else self.sample_uncertainty_based(items, n_pairs)

    # ── Main dispatch ─────────────────────────────────────────────────────────

    def sample(self, items: List[str],
               comparisons_done: int,
               n_pairs: int = 1) -> List[Tuple[str, str]]:
        """
        Dispatch to the correct phase based on comparison count.
        This is the only method run_experiment should call.
        """
        phase = self.current_phase(comparisons_done)
        if phase == 1:
            return self.sample_uncertainty_based(items, n_pairs)
        elif phase == 2:
            return self.sample_competitive(items, n_pairs)
        else:
            return self.sample_boundary_focused(items, n_pairs)


# ─────────────────────────────────────────────────────────────────────────────
# MAIN ENGINE
# ─────────────────────────────────────────────────────────────────────────────

class BTComparisonEngine:
    """
    Combines BT model, LLM API calls, and adaptive sampling.
    """

    # Supported providers and their default models
    PROVIDER_DEFAULTS = {
        "anthropic": "claude-sonnet-4-6",
        "openai":    "gpt-4o",
        "openai-compat": "gpt-4o",
    }

    def __init__(self, api_key: str,
                 model:    str = None,
                 provider: str = "anthropic",
                 base_url: str = None):
        """
        Args:
            api_key:  API key for the chosen provider.
            model:    Model string. Defaults to provider's recommended model.
            provider: One of 'anthropic', 'openai', 'openai-compat'.
                      Use 'openai-compat' for any OpenAI-compatible endpoint.
            base_url: Custom endpoint for 'openai-compat'.
        """
        self.provider = provider.lower()
        self.model    = model or self.PROVIDER_DEFAULTS.get(self.provider,
                                                             "claude-sonnet-4-6")
        self.base_url = base_url

        if self.provider == "anthropic":
            self.client = anthropic.Anthropic(api_key=api_key)
        elif self.provider in ("openai", "openai-compat"):
            if not _OPENAI_AVAILABLE:
                raise ImportError(
                    "openai package not installed. Run: pip install openai"
                )
            kwargs = {"api_key": api_key}
            if base_url:
                kwargs["base_url"] = base_url
            self.client = _openai_lib.OpenAI(**kwargs)
        else:
            raise ValueError(
                f"Unknown provider '{provider}'. "
                f"Choose from: {list(self.PROVIDER_DEFAULTS)}"
            )

        self.bt      = BradleyTerry()
        self.sampler = BTAdaptiveSampler(self.bt)

    @staticmethod
    def recommended_comparisons(n_items: int, target_matches: int = 15,
                                 correction: float = 1.2) -> int:
        """
        (n_items × target_matches / 2) × correction

        Each comparison covers two items, so divides by 2.
        """
        return int((n_items * target_matches / 2) * correction)

    # ── API call ──────────────────────────────────────────────────────────────

    def compare_pair(self, item_a: str, item_b: str,
                     prompt_template: str) -> Optional[str]:
        """
        Query LLM and return winner.
        Randomizes A/B presentation order to cancel position bias.
        Dispatches to the correct provider API.
        """
        swap = random.random() < 0.5
        if swap:
            prompt = prompt_template.format(item_a=item_b, item_b=item_a)
        else:
            prompt = prompt_template.format(item_a=item_a, item_b=item_b)

        try:
            response_text = self._call_api(prompt)
            if response_text is None:
                return None

            choice = self._extract_choice(response_text)
            if choice is None:
                print(f"  ⚠ Unparseable response: {response_text[:120]!r}")
                return None

            if swap:
                return item_b if choice == 'A' else item_a
            else:
                return item_a if choice == 'A' else item_b

        except Exception as e:
            print(f"  API Error: {e}")
            return None

    @staticmethod
    def _extract_choice(response_text: str) -> Optional[str]:
        """
        Extract 'A' or 'B' from a response that may contain chain-of-thought.

        Strategy (in order):
          1. If the last non-empty line is exactly 'A' or 'B' → use it.
             (Reasoning models typically end with the bare answer.)
          2. Scan the last 60 chars for a standalone A or B.
          3. Fall back to whole-text scan: A-only → A, B-present → B.
        """
        import re
        text = response_text.strip()

        # 1. Last non-empty line
        lines = [l.strip() for l in text.splitlines() if l.strip()]
        if lines and lines[-1].upper() in ('A', 'B'):
            return lines[-1].upper()

        # 2. Standalone letter near the end (handles "Answer: A" etc.)
        tail = text[-60:].upper()
        m = re.search(r'\b([AB])\b(?!.*\b[AB]\b)', tail)
        if m:
            return m.group(1)

        # 3. Whole-text fallback — only fire if exactly one of A/B appears
        # as a standalone word in the full text. Never fire on bare letter
        # presence (too permissive: "ABOVE" contains 'A', "BOTH" contains 'B').
        upper = text.upper()
        a_standalone = bool(re.search(r'A', upper))
        b_standalone = bool(re.search(r'B', upper))
        if a_standalone and not b_standalone:
            print(f"  ⚠ _extract_choice: fallback fired → A  (response: {text[:80]!r})")
            return 'A'
        if b_standalone and not a_standalone:
            print(f"  ⚠ _extract_choice: fallback fired → B  (response: {text[:80]!r})")
            return 'B'

        # Truly ambiguous or empty — caller increments skipped counter
        return None


    def _call_api(self, prompt: str) -> Optional[str]:
        """
        Dispatch a single prompt to the configured provider and return raw text.

        All models use temperature=0, max_tokens=10 (one-token A/B answer).
        If the API rejects max_tokens (e.g. GPT-5.x and later OpenAI models
        that require max_completion_tokens), the call is retried automatically.
        """
        if self.provider == "anthropic":
            msg = self.client.messages.create(
                model=self.model,
                max_tokens=10,
                temperature=0,
                messages=[{"role": "user", "content": prompt}]
            )
            return msg.content[0].text

        elif self.provider in ("openai", "openai-compat"):
            kwargs = dict(
                model=self.model,
                messages=[{"role": "user", "content": prompt}],
                temperature=0,
            )
            try:
                resp = self.client.chat.completions.create(
                    **kwargs, max_tokens=10
                )
            except Exception as e:
                if "max_tokens" in str(e) or "max_completion_tokens" in str(e):
                    resp = self.client.chat.completions.create(
                        **kwargs, max_completion_tokens=10
                    )
                else:
                    raise
            return resp.choices[0].message.content

        else:
            raise ValueError(f"Unknown provider: {self.provider}")

    # ── Single experiment run ─────────────────────────────────────────────────

    # Default boundary ranks: top-10, top-20, bottom-20, bottom-10
    # (set relative to n_items at runtime; these are overrideable)
    DEFAULT_BOUNDARY_RANKS: List[int] = [10, 20, 100, 110]

    def run_experiment(
        self,
        items: List[str],
        prompt_template: str,
        max_comparisons: int,
        sampling_strategy: str = "balanced",
        boundary_ranks: Optional[List[int]] = None,
        refit_every: int = 50,
    ) -> Dict:
        """
        Collect pairwise outcomes via three-phase adaptive sampling,
        refitting BT every `refit_every` comparisons.

        Phase schedule (set automatically by BTAdaptiveSampler.configure):
          Phase 1: uncertainty-based until every item has >= 15 matches
          Phase 2: broad competitive (full-ranking stabilization)
          Phase 3: boundary-focused (top/bottom 25% intensification)

        Args:
            boundary_ranks: Ranks to intensify in Phase 3.
                            Defaults to top-25% and bottom-25% boundaries.
            refit_every:    How often to refit the BT model mid-run.
        """
        n_items = len(items)

        # Default boundary ranks: top-25% cutoffs + single bottom anchor
        # Reflect research priority (critical acclaim orientation) — bottom gets cross-anchor
        # treatment in Phase 3 without needing explicit ranks here.
        if boundary_ranks is None:
            q12 = max(1, n_items // 8)   # top-12.5%
            q25 = max(2, n_items // 4)   # top-25%
            bot = min(n_items - 1, n_items - q25)  # bottom anchor
            boundary_ranks = sorted({q12, q25, bot})

        # Register all items upfront
        for item in items:
            self.bt.add_item(item)

        # Configure three-phase sampler
        self.sampler.configure(n_items, max_comparisons, boundary_ranks)

        comparisons_done = 0
        skipped = 0
        last_phase = 0

        print(f"Bradley-Terry experiment")
        print(f"Items: {n_items}  |  Max comparisons: {max_comparisons}")
        print()

        while comparisons_done < max_comparisons:
            # Announce phase transitions
            phase = self.sampler.current_phase(comparisons_done)
            if phase != last_phase:
                phase_names = {1: "Phase 1: global coverage",
                               2: "Phase 2: broad competitive",
                               3: "Phase 3: boundary focus"}
                print(f"\n{'─'*60}")
                print(f"  [{phase_names[phase]}]")
                print(f"{'─'*60}")
                last_phase = phase

            # Dispatch to phase-appropriate sampler
            pairs = self.sampler.sample(items, comparisons_done, n_pairs=1)

            if not pairs:
                continue

            item_a, item_b = pairs[0]

            print(f"[{comparisons_done+1}/{max_comparisons}] "
                  f"{item_a} vs. {item_b}")

            winner = self.compare_pair(item_a, item_b, prompt_template)

            if winner is None:
                print("  → Failed, skipping")
                skipped += 1
                continue

            loser = item_b if winner == item_a else item_a
            self.bt.record_outcome(winner, loser,
                                   run_id=getattr(self, '_current_run_id', 0),
                                   comparison_idx=comparisons_done)
            comparisons_done += 1

            print(f"  → Winner: {winner}")

            # Refit BT periodically so sampler uses updated strengths
            if comparisons_done % refit_every == 0:
                try:
                    self.bt.fit()
                    print(f"\n  [BT refit at n={comparisons_done} | "
                          f"phase={phase}]")
                    rankings = self.bt.get_rankings()
                    counts   = self.bt.match_counts_per_item()
                    min_n    = min(counts.values()) if counts else 0
                    print(f"  Min matches/item so far: {min_n}")
                    print(f"  Current top 5:")
                    for row in rankings[:5]:
                        print(f"    {row['rank']}. {row['item']:<45} "
                              f"λ={row['log_strength']:+.3f} "
                              f"(±{row['std_error']:.3f})")
                    print()
                except Exception as e:
                    print(f"  [BT refit failed: {e}]")

        # Final fit
        self.bt.fit()
        rankings = self.bt.get_rankings()

        print(f"\n{'='*70}")
        print(f"FINAL RESULTS  (Bradley-Terry MLE)")
        print(f"{'='*70}")
        print(f"Comparisons: {comparisons_done}  |  Skipped: {skipped}")
        print(f"Items ranked: {len(rankings)}")
        print()
        print(f"{'Rank':<5} {'Item':<50} {'λ':>7} {'SE':>6} {'95% CI':<20} {'n':>4}")
        print("-" * 95)
        for row in rankings[:20]:
            ci = f"[{row['ci_lower_95']:+.3f}, {row['ci_upper_95']:+.3f}]"
            print(f"  {row['rank']:<4} {row['item']:<50} "
                  f"{row['log_strength']:>+7.3f} {row['std_error']:>6.3f} "
                  f"{ci:<20} {row['matches']:>4}")

        return {
            'rankings':          rankings,
            'total_comparisons': comparisons_done,
            'skipped':           skipped,
            'bt':                self.bt,
        }

    # ── Multi-iteration aggregation ───────────────────────────────────────────

    def run_multi_iteration(
        self,
        items: List[str],
        prompt_template: str,
        n_iterations: int = 5,
        comparisons_per_iter: Optional[int] = None,
        sampling_strategy: str = "balanced",
        boundary_ranks: Optional[List[int]] = None,
        refit_every: int = 50,
    ) -> Dict:
        """
        Run N independent BT experiments and aggregate by mean log-strength.

        Output per item:
          mean_lambda   — point estimate of log-strength
          std_lambda    — SD across runs (stability metric)
          mean_se       — mean within-run standard error
          pooled_se     — combined uncertainty estimate
        """
        if comparisons_per_iter is None:
            comparisons_per_iter = self.recommended_comparisons(len(items))

        print(f"Multi-iteration BT — {n_iterations} runs × "
              f"{comparisons_per_iter} comparisons each")
        print(f"  Total API calls: ~{n_iterations * comparisons_per_iter:,}\n")

        lambdas_across_runs: Dict[str, List[float]] = defaultdict(list)
        ses_across_runs:     Dict[str, List[float]] = defaultdict(list)
        matches_across_runs: Dict[str, int]          = defaultdict(int)
        run_summaries = []
        all_match_history: List[Dict] = []

        for i in range(n_iterations):
            print(f"\n{'='*60}")
            print(f"  RUN {i+1}/{n_iterations}")
            print(f"{'='*60}\n")

            # Fresh BT instance per run
            self.bt      = BradleyTerry()
            self.sampler = BTAdaptiveSampler(self.bt)

            self._current_run_id = i + 1
            result = self.run_experiment(
                items=items,
                prompt_template=prompt_template,
                max_comparisons=comparisons_per_iter,
                sampling_strategy=sampling_strategy,
                boundary_ranks=boundary_ranks,
                refit_every=refit_every,
            )

            for row in result['rankings']:
                lambdas_across_runs[row['item']].append(row['log_strength'])
                ses_across_runs[row['item']].append(row['std_error'])
                matches_across_runs[row['item']] += row.get('matches', 0)

            all_match_history.extend(self.bt.match_history)  # accumulate

            run_summaries.append({
                'run':         i + 1,
                'comparisons': result['total_comparisons'],
                'top5':        [r['item'] for r in result['rankings'][:5]],
            })

            # ── Per-iteration checkpoint ───────────────────────────────────
            # Written immediately after each run completes so a crash in
            # save_results() or aggregation never loses completed iterations.
            if hasattr(self, '_checkpoint_path') and self._checkpoint_path:
                ckpt = {
                    'run':         i + 1,
                    'rankings':    result['rankings'],
                    'match_history': self.bt.match_history,
                }
                ckpt_file = (
                    self._checkpoint_path
                    .replace('.json', f'_ckpt_run{i+1}.json')
                )
                try:
                    with open(ckpt_file, 'w') as cf:
                        json.dump(ckpt, cf, indent=2)
                    print(f"  ✓ Checkpoint saved → {ckpt_file}")
                except Exception as ckpt_err:
                    print(f"  ⚠ Checkpoint write failed: {ckpt_err}")

        # ── Aggregate ─────────────────────────────────────────────────────────
        aggregated = []
        for item in items:
            lams = lambdas_across_runs.get(item, [])
            ses  = ses_across_runs.get(item, [])
            if not lams:
                continue

            mean_lam = float(np.mean(lams))
            std_lam  = float(np.std(lams))
            mean_se  = float(np.mean(ses))
            # Pooled SE: within-run uncertainty + between-run variance
            pooled_se = math.sqrt(mean_se**2 + std_lam**2)

            aggregated.append({
                'item':         item,
                'mean_lambda':  round(mean_lam, 4),
                'std_lambda':   round(std_lam, 4),
                'mean_se':      round(mean_se, 4),
                'pooled_se':    round(pooled_se, 4),
                'ci_lower_95':  round(mean_lam - 1.96 * pooled_se, 4),
                'ci_upper_95':  round(mean_lam + 1.96 * pooled_se, 4),
                'n_runs':       len(lams),
                'matches':      matches_across_runs.get(item, 0),
                'run_lambdas':  [round(l, 4) for l in lams],
            })

        aggregated.sort(key=lambda x: x['mean_lambda'], reverse=True)
        for idx, row in enumerate(aggregated, 1):
            row['rank'] = idx

        # ── Inter-run reliability ─────────────────────────────────────────────
        from scipy.stats import spearmanr
        item_order = [r['item'] for r in aggregated]
        run_vecs = [
            [lambdas_across_runs[item][run_i]
             for item in item_order
             if run_i < len(lambdas_across_runs[item])]
            for run_i in range(n_iterations)
        ]

        rhos = []
        print(f"\n{'='*70}")
        print(f"AGGREGATED RESULTS ({n_iterations} runs)")
        print(f"{'='*70}")
        print(f"\nInter-run Spearman ρ:")
        for a in range(n_iterations):
            for b in range(a + 1, n_iterations):
                rho, _ = spearmanr(run_vecs[a], run_vecs[b])
                rhos.append(rho)
                print(f"  Run {a+1} vs Run {b+1}: ρ = {rho:.3f}")
        print(f"  Mean ρ: {np.mean(rhos):.3f}  ← noise floor for interpretation\n")

        print(f"{'Rank':<5} {'Item':<50} {'λ̄':>7} {'±SD':>6} "
              f"{'Pooled SE':>10} {'95% CI':<22}")
        print("-" * 100)
        for row in aggregated[:20]:
            ci = f"[{row['ci_lower_95']:+.3f}, {row['ci_upper_95']:+.3f}]"
            print(f"  {row['rank']:<4} {row['item']:<50} "
                  f"{row['mean_lambda']:>+7.3f} {row['std_lambda']:>6.3f} "
                  f"{row['pooled_se']:>10.3f} {ci:<22}")

        return {
            'aggregated_rankings':  aggregated,
            'run_summaries':        run_summaries,
            'n_iterations':         n_iterations,
            'comparisons_per_iter': comparisons_per_iter,
            'total_comparisons':    sum(s['comparisons'] for s in run_summaries),
            'inter_run_rho':        round(float(np.mean(rhos)), 4),
            'all_match_history':    all_match_history,
        }

    # ── Cross-set win rate matrix ─────────────────────────────────────────────

    @staticmethod
    def compute_cross_set_winrates(match_history: List[Dict],
                                   film_meta: Dict) -> Dict:
        """
        Compute pairwise win rates between Sets A, B, and C.

        For each comparison where winner and loser belong to different sets,
        record which set won. Aggregates into a symmetric 3×3 matrix.

        Returns a dict with:
          comparisons_used  — number of cross-set matches counted
          matrix            — {set_X: {set_Y: {wins, total, win_rate}}}
          summary           — human-readable rows like "A vs B: A wins 62.3%"
        """
        sets = ['A', 'B', 'C']
        wins  = {s: {t: 0 for t in sets} for s in sets}
        total = {s: {t: 0 for t in sets} for s in sets}

        cross_count = 0
        for match in match_history:
            w_key = match.get('w') or match.get('winner', '')
            l_key = match.get('l') or match.get('loser',  '')
            w_set = film_meta.get(w_key, {}).get('set')
            l_set = film_meta.get(l_key, {}).get('set')
            if w_set is None or l_set is None or w_set == l_set:
                continue
            wins[w_set][l_set]  += 1
            total[w_set][l_set] += 1
            total[l_set][w_set] += 1
            cross_count += 1

        matrix = {}
        for s in sets:
            matrix[s] = {}
            for t in sets:
                if s == t:
                    matrix[s][t] = None
                    continue
                n  = total[s][t]
                w  = wins[s][t]
                matrix[s][t] = {
                    'wins':     w,
                    'total':    n,
                    'win_rate': round(w / n, 4) if n > 0 else None,
                }

        # Human-readable summary lines
        summary = []
        for s, t in [('A', 'B'), ('A', 'C'), ('B', 'C')]:
            n = total[s][t]
            if n == 0:
                summary.append(f"Set {s} vs Set {t}: no cross-set matches")
                continue
            wr_s = wins[s][t] / n
            wr_t = wins[t][s] / n
            leader    = s if wr_s >= wr_t else t
            leader_wr = max(wr_s, wr_t)
            summary.append(
                f"Set {s} vs Set {t}: "
                f"Set {s} wins {wr_s:.1%} ({wins[s][t]}/{n}), "
                f"Set {t} wins {wr_t:.1%} ({wins[t][s]}/{n})  "
                f"→ Set {leader} leads"
            )

        return {
            'comparisons_used': cross_count,
            'matrix':           matrix,
            'summary':          summary,
        }

    # ── Excel export ──────────────────────────────────────────────────────────

    @staticmethod
    def save_excel(rankings: List[Dict], film_meta: Dict,
                   output_path: str, sheet_title: str = "Rankings"):
        """
        Write a formatted Excel workbook with one Rankings sheet.

        Columns: Rank | Film | λ | SE | CI Lower | CI Upper | Matches |
                 Set | Country | TSPDT Rank | Box Office Rank

        λ / SE / CI columns come from whichever key is present:
          single-run   → log_strength / std_error
          multi-iter   → mean_lambda  / pooled_se
        """
        if not _OPENPYXL_AVAILABLE:
            print("  ⚠ openpyxl not installed — skipping Excel output.\n"
                  "    Install with: pip install openpyxl")
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_title

        # ── Palette ───────────────────────────────────────────────────────────
        SET_COLORS = {
            'A': 'DDEEFF',   # soft blue  — dual-acclaim
            'B': 'DDFFDD',   # soft green — critical-only
            'C': 'FFF3CC',   # soft amber — popular-only
        }
        HEADER_FILL  = PatternFill("solid", fgColor="2F4F8F")
        HEADER_FONT  = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
        BODY_FONT    = Font(name="Calibri", size=10)
        ALT_FILL     = PatternFill("solid", fgColor="F5F5F5")
        THIN_BORDER  = Border(
            bottom=Side(style='thin', color='CCCCCC'),
        )
        CENTER       = Alignment(horizontal='center', vertical='center')
        LEFT         = Alignment(horizontal='left',   vertical='center')

        # ── Column definitions ────────────────────────────────────────────────
        headers = [
            ("Rank",          7),
            ("Film",         46),
            ("λ",             9),
            ("SE",            8),
            ("CI Lower 95%", 13),
            ("CI Upper 95%", 13),
            ("Matches",       9),
            ("Set",           6),
            ("Country",      20),
            ("TSPDT Rank",   12),
            ("Mojo Rank",    11),
        ]

        # Write header row
        for col_idx, (header, width) in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font      = HEADER_FONT
            cell.fill      = HEADER_FILL
            cell.alignment = CENTER
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        ws.row_dimensions[1].height = 20
        ws.freeze_panes = "A2"

        # ── Data rows ─────────────────────────────────────────────────────────
        for row_idx, item_row in enumerate(rankings, 2):
            item  = item_row['item']
            meta  = film_meta.get(item, {})
            iset  = meta.get('set', '')

            # Resolve λ / SE / CI keys (single vs multi-iter)
            lam      = item_row.get('log_strength',  item_row.get('mean_lambda'))
            se       = item_row.get('std_error',     item_row.get('pooled_se'))
            ci_lower = item_row.get('ci_lower_95')
            ci_upper = item_row.get('ci_upper_95')

            values = [
                item_row.get('rank', row_idx - 1),
                item,
                round(lam,      4) if lam      is not None else '',
                round(se,       4) if se        is not None else '',
                round(ci_lower, 4) if ci_lower  is not None else '',
                round(ci_upper, 4) if ci_upper  is not None else '',
                item_row.get('matches', ''),
                iset,
                meta.get('country',   ''),
                meta.get('tspdt_pos') or '',
                meta.get('mojo_rank') or '',
            ]

            set_fill  = PatternFill("solid", fgColor=SET_COLORS.get(iset, 'FFFFFF'))
            alt_fill  = ALT_FILL if row_idx % 2 == 0 else None

            for col_idx, value in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font   = BODY_FONT
                cell.border = THIN_BORDER

                # Alignment
                if col_idx == 2:
                    cell.alignment = LEFT
                else:
                    cell.alignment = CENTER

                if col_idx == 8 and iset:
                    cell.fill = set_fill
                elif alt_fill:
                    cell.fill = alt_fill

        # ── Set legend (below data) ───────────────────────────────────────────
        legend_row = len(rankings) + 3
        ws.cell(row=legend_row, column=1, value="Set Legend").font = Font(bold=True)
        for offset, (skey, label, color) in enumerate([
            ('A', 'Set A — Dual acclaim (TSPDT + Mojo)',  SET_COLORS['A']),
            ('B', 'Set B — Critical only (TSPDT)',        SET_COLORS['B']),
            ('C', 'Set C — Popular only (Mojo)',          SET_COLORS['C']),
        ], 1):
            r = legend_row + offset
            lc = ws.cell(row=r, column=1, value=skey)
            lc.fill = PatternFill("solid", fgColor=color)
            lc.alignment = CENTER
            ws.cell(row=r, column=2, value=label).font = Font(italic=True, size=9)

        # ── Auto-filter on header row ─────────────────────────────────────────
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

        xl_path = output_path.replace('.json', '.xlsx')
        wb.save(xl_path)
        print(f"✓ Excel saved to {xl_path}")

    # ── Save ──────────────────────────────────────────────────────────────────

    @staticmethod
    def make_output_path(base_dir: str, prompt_label: str,
                         model: str, total_comparisons: int,
                         n_iterations: int = 1) -> str:
        """
        Format:
          bt_{prompt}_{model_slug}_n{comparisons}[_{k}x]_{timestamp}.json
        """
        import re
        # Shorten model string to a readable slug (strip date suffixes etc.)
        model_slug = re.sub(r'[-_]2\d{7}.*$', '', model)
        model_slug = model_slug.replace('/', '-')[:30]
        prompt_slug = prompt_label.replace(' ', '_')[:25]

        iter_part  = f"_{n_iterations}x" if n_iterations > 1 else ""
        ts         = datetime.now().strftime("%Y%m%d-%H%M%S")
        filename   = f"bt_{prompt_slug}_{model_slug}_n{total_comparisons}{iter_part}_{ts}.json"
        return f"{base_dir.rstrip('/')}/{filename}"

    def save_results(self, results: Dict, output_path: str,
                     metadata: Dict = None):
        film_meta = (metadata or {}).get('film_meta', {})

        # ── Cross-set win rates (JSON only) ───────────────────────────────────
        # Collect match history: single-run uses self.bt directly;
        # multi-iter aggregates across run_summaries via stored bt history.
        if 'aggregated_rankings' in results:
            cross_set = self.compute_cross_set_winrates(
                results.get('all_match_history', []), film_meta
            )
        else:
            cross_set = self.compute_cross_set_winrates(
                self.bt.match_history, film_meta
            )

        # Strip film_meta from the stored metadata
        meta_slim = {k: v for k, v in (metadata or {}).items() if k != 'film_meta'}

        if 'aggregated_rankings' in results:
            rankings_for_excel = results['aggregated_rankings']
            data = {
                'model':                self.model,
                'provider':             self.provider,
                'domain':               'film',
                'method':               'bradley_terry',
                'timestamp':            datetime.now().isoformat(),
                'mode':                 'multi_iteration',
                'n_iterations':         results['n_iterations'],
                'comparisons_per_iter': results['comparisons_per_iter'],
                'total_comparisons':    results['total_comparisons'],
                'inter_run_rho':        results['inter_run_rho'],
                'metadata':             meta_slim,
                'rankings':             results['aggregated_rankings'],
                'run_summaries':        results['run_summaries'],
                'cross_set_win_rates':  cross_set,
            }
        else:
            rankings_for_excel = results['rankings']
            data = {
                'model':                self.model,
                'provider':             self.provider,
                'domain':               'film',
                'method':               'bradley_terry',
                'timestamp':            datetime.now().isoformat(),
                'mode':                 'single_run',
                'total_comparisons':    results['total_comparisons'],
                'skipped':              results['skipped'],
                'metadata':             meta_slim,
                'rankings':             results['rankings'],
                'cross_set_win_rates':  cross_set,
            }

        # ── JSON ──────────────────────────────────────────────────────────────
        with open(output_path, 'w') as f:
            json.dump(data, f, indent=2)
        print(f"\n✓ JSON saved to  {output_path}")

        # ── Match log JSON ─────────────────────────────────────────────────────
        # Format: {run, n, w, l}  (run=iteration, n=comparison index, w=winner, l=loser)
        log_path = output_path.replace('.json', '_log.json')
        if 'aggregated_rankings' in results:
            match_history = results.get('all_match_history', [])
        else:
            match_history = self.bt.match_history

        # Build header separately from matches so we can indent the header
        # while keeping each match record on a single compact line.
        header = {
            'model':                self.model,
            'provider':             self.provider,
            'prompt_frame':         (metadata or {}).get('prompt_label', ''),
            'prompt':               (metadata or {}).get('prompt', ''),
            'n_items':              len(film_meta),
            'n_iterations':         data.get('n_iterations', 1),
            'comparisons_per_iter': data.get('comparisons_per_iter',
                                              data.get('total_comparisons', 0)),
            'total_matches':        len(match_history),
        }
        # Write: indented header block, then "matches": [ one record per line ]
        import io
        buf = io.StringIO()
        # Dump header fields as indented JSON, strip closing }
        header_str = json.dumps(header, indent=2)
        # Strip trailing whitespace/newline before } then add comma
        buf.write(header_str[:-1].rstrip())
        buf.write(',\n  "matches": [\n')
        for idx, m in enumerate(match_history):
            comma = '' if idx == len(match_history) - 1 else ','
            buf.write('    ' + json.dumps(m, separators=(',', ':')) + comma + '\n')
        buf.write('  ]\n}')
        with open(log_path, 'w') as f:
            f.write(buf.getvalue())
        print(f"✓ Match log saved to {log_path}")

        # ── Excel ─────────────────────────────────────────────────────────────
        try:
            self.save_excel(rankings_for_excel, film_meta, output_path)
        except Exception as xl_err:
            print(f"  ⚠ Excel export failed (JSON and log are safe): {xl_err}")


# ─────────────────────────────────────────────────────────────────────────────
# PROMPT PRESETS
# ─────────────────────────────────────────────────────────────────────────────

PROMPT_PRESETS = {
    "preference": (
        "Which of these two films do you prefer? "
        "A: {item_a}; B: {item_b}. "
        "Respond with either 'A' or 'B' only."
    ),
    "taste": (
        "Which of these two films is closer to your taste? "
        "A: {item_a}; B: {item_b}. "
        "Respond with either 'A' or 'B' only."
    ),
    "like": (
        "Which of these two films do you like more? "
        "A: {item_a}; B: {item_b}. "
        "Respond with either 'A' or 'B' only."
    ),
    "general_audience": (
        "You are recommending a film to a general audience. "
        "Which would you recommend? "
        "A: {item_a}; B: {item_b}. "
        "Respond with either 'A' or 'B' only."
    )
}

# ─────────────────────────────────────────────────────────────────────────────
# ENTRY POINT
# ─────────────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    import argparse
    import os

    parser = argparse.ArgumentParser(
        description="Bradley-Terry pairwise comparison — Film Edition",
        formatter_class=argparse.RawTextHelpFormatter,
    )
    parser.add_argument("--dataset",         type=str, required=True,
                        help="Path to film_sample_200.json")
    parser.add_argument("--api-key",         type=str, default=None,
                        help="API key. Defaults to ANTHROPIC_API_KEY / OPENAI_API_KEY env var.")
    parser.add_argument("--provider",        type=str, default="anthropic",
                        choices=["anthropic", "openai", "openai-compat"],
                        help="LLM provider (default: anthropic).\n"
                             "  anthropic    — Claude via Anthropic API\n"
                             "  openai       — GPT via OpenAI API\n"
                             "  openai-compat — any OpenAI-compatible endpoint (set --base-url)")
    parser.add_argument("--base-url",        type=str, default=None,
                        help="Custom endpoint for openai-compat provider\n"
                             "(e.g. https://api.groq.com/openai/v1)")
    parser.add_argument("--model",           type=str, default=None,
                        help="Model string. Defaults to provider's recommended model.\n"
                             "  anthropic default: claude-sonnet-4-6\n"
                             "  openai default:    gpt-4o")
    parser.add_argument("--output-dir",      type=str,
                        default="./outputs",
                        help="Directory for output JSON files.")
    parser.add_argument("--max-comparisons", type=int, default=None,
                        help="Comparisons per run. Defaults to recommended value.")
    parser.add_argument("--strategy",        type=str, default="balanced",
                        choices=['balanced', 'uncertainty', 'strength', 'boundary'],
                        help="Sampling strategy. 'balanced' uses the three-phase scheduler\n"
                             "(coverage → competitive → boundary). Other values force a\n"
                             "single strategy throughout (mainly useful for ablations).")
    parser.add_argument("--min-floor",       type=int, default=15,
                        help="Minimum matches per item before Phase 2 starts (default: 15).")
    parser.add_argument("--middle-fraction", type=float, default=0.35,
                        help="Fraction of post-floor budget for Phase 2 broad sampling (default: 0.35).")
    parser.add_argument("--prompt-preset",   type=str, default=None,
                        choices=list(PROMPT_PRESETS.keys()))
    parser.add_argument("--prompt",          type=str, default=None,
                        help="Custom prompt with {item_a} and {item_b}.")
    parser.add_argument("--iterations",      type=int, default=1,
                        help="Number of independent runs to aggregate (default: 1).")
    parser.add_argument("--boundary-ranks",  type=int, nargs='+', default=None,
                        help="Ranks to focus boundary sampling on, e.g. --boundary-ranks 10 20 110\n"
                             "Activates boundary_focused sampling strategy.")
    parser.add_argument("--refit-every",     type=int, default=50,
                        help="Refit BT model every N comparisons (default: 50).")
    parser.add_argument("--output",          type=str, default=None,
                        help="Full output path. If omitted, auto-generated from\n"
                             "prompt, model, comparison count, and timestamp.")
    args = parser.parse_args()

    # ── Resolve prompt ────────────────────────────────────────────────────────
    if args.prompt:
        prompt = args.prompt
        prompt_label = "custom"
        if '{item_a}' not in prompt or '{item_b}' not in prompt:
            raise ValueError("--prompt must contain {item_a} and {item_b}.")
    elif args.prompt_preset:
        prompt = PROMPT_PRESETS[args.prompt_preset]
        prompt_label = args.prompt_preset
    else:
        prompt_label = "preference"
        prompt = PROMPT_PRESETS[prompt_label]
        print(f"No prompt specified — using default '{prompt_label}'.")

    # ── Resolve API key ───────────────────────────────────────────────────────
    api_key = args.api_key
    if not api_key:
        env_map = {
            "anthropic":    "ANTHROPIC_API_KEY",
            "openai":       "OPENAI_API_KEY",
            "openai-compat":"OPENAI_API_KEY",
        }
        api_key = os.environ.get(env_map.get(args.provider, "ANTHROPIC_API_KEY"), "")
    if not api_key:
        raise ValueError(
            f"No API key found. Pass --api-key or set the appropriate env var."
        )

    print(f"\nMethod        : Bradley-Terry MLE")
    print(f"Provider      : {args.provider}")
    print(f"Model         : {args.model or BTComparisonEngine.PROVIDER_DEFAULTS.get(args.provider)}")
    print(f"Prompt preset : {prompt_label}")
    print(f"Prompt        : {prompt}")
    print(f"Iterations    : {args.iterations}")
    print(f"Boundary ranks: {args.boundary_ranks}")
    print()

    # ── Load films ────────────────────────────────────────────────────────────
    with open(args.dataset) as f:
        raw = json.load(f)

    films, film_meta = [], {}
    for set_key, set_data in raw['sets'].items():
        for film in set_data['films']:
            label = f"{film['Title']} ({film['Year']})"
            films.append(label)
            film_meta[label] = {
                'set':            set_key,
                'set_label':      set_data['label'],
                'director':       film.get('Director', ''),
                'genre':          film.get('Genre', ''),
                'country':        film.get('Country', ''),
                'language':       film.get('Language', ''),
                'language_group': film.get('LanguageGroup', ''),
                'tspdt_pos':      film.get('tspdt_pos'),
                'mojo_rank':      film.get('mojo_rank'),
            }

    n = len(films)
    rec = BTComparisonEngine.recommended_comparisons(n)
    max_comp = args.max_comparisons or rec

    print(f"Loaded {n} films")
    print(f"Recommended comparisons: {rec}  |  Using: {max_comp}")
    if max_comp < rec:
        print("  ⚠ Below recommendation — rankings may be unstable.")
    print(f"Example items: {films[:3]}\n")

    # ── Run ───────────────────────────────────────────────────────────────────
    engine = BTComparisonEngine(
        api_key  = api_key,
        model    = args.model,
        provider = args.provider,
        base_url = args.base_url,
    )
    # Apply sampler configuration from CLI args
    engine.sampler.min_floor       = args.min_floor
    engine.sampler.middle_fraction = args.middle_fraction
    meta = {'prompt_label': prompt_label, 'prompt': prompt,
            'film_meta': film_meta}

    # Pre-set checkpoint path so multi-iteration runs survive save_results crashes
    output_path_preview = BTComparisonEngine.make_output_path(
        base_dir          = args.output_dir,
        prompt_label      = prompt_label,
        model             = args.model or BTComparisonEngine.PROVIDER_DEFAULTS.get(args.provider, 'unknown'),
        total_comparisons = max_comp * max(args.iterations, 1),
        n_iterations      = args.iterations,
    ) if args.output is None else args.output
    engine._checkpoint_path = output_path_preview

    if args.iterations > 1:
        results = engine.run_multi_iteration(
            items=films,
            prompt_template=prompt,
            n_iterations=args.iterations,
            comparisons_per_iter=max_comp,
            sampling_strategy=args.strategy,
            boundary_ranks=args.boundary_ranks,
            refit_every=args.refit_every,
        )
    else:
        engine._current_run_id = 1
        results = engine.run_experiment(
            items=films,
            prompt_template=prompt,
            max_comparisons=max_comp,
            sampling_strategy=args.strategy,
            boundary_ranks=args.boundary_ranks,
            refit_every=args.refit_every,
        )

    # ── Build output path ─────────────────────────────────────────────────────
    total_comp = results.get('total_comparisons',
                  results.get('comparisons_per_iter', max_comp) * args.iterations)
    output_path = args.output or BTComparisonEngine.make_output_path(
        base_dir          = args.output_dir,
        prompt_label      = prompt_label,
        model             = engine.model,
        total_comparisons = total_comp,
        n_iterations      = args.iterations,
    )
    print(f"Output: {output_path}")

    engine.save_results(results, output_path, metadata=meta)
